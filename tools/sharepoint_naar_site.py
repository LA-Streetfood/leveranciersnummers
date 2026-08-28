#!/usr/bin/env python3
"""
Haalt 'NP - Leveranciers -Artikelnummers.xlsx' uit SharePoint op en zet de
inhoud om naar leveranciers-data.json, het bestand dat de website uitleest.

Draait normaal via .github/workflows/leveranciersnummers.yml. Handmatig testen
op een gedownload bestand kan ook:

    python3 tools/sharepoint_naar_site.py --bestand ~/Downloads/lijst.xlsx --droog

Nodig als omgevingsvariabelen (in GitHub als Secrets):
    SP_TENANT_ID, SP_CLIENT_ID, SP_CLIENT_SECRET
"""

import argparse
import io
import json
import os
import re
import sys
import urllib.parse
from datetime import datetime, timezone

import requests
from openpyxl import load_workbook

HOSTNAME = os.environ.get("SP_HOSTNAME", "latinamericanfoodbv.sharepoint.com")
SITE_PATH = os.environ.get("SP_SITE_PATH", "/sites/kantoor")
FILE_PATH = os.environ.get("SP_FILE_PATH", "Artikellijsten/NP - Leveranciers -Artikelnummers.xlsx")

# Volgorde van de productkolommen in de tabel op de website.
CODES = ["LA152", "LA824", "LA768", "LA823", "LA7600", "LA882", "LA979", "LA115"]

LANDEN = {
    "NL": ("nl", "nederland", "nederlandse", "holland"),
    "BE": ("be", "belgie", "belgië", "belgische", "belgique", "belgium"),
}


def log(bericht):
    print(bericht, flush=True)


def graph_token():
    tenant = os.environ["SP_TENANT_ID"]
    antwoord = requests.post(
        "https://login.microsoftonline.com/%s/oauth2/v2.0/token" % tenant,
        data={
            "client_id": os.environ["SP_CLIENT_ID"],
            "client_secret": os.environ["SP_CLIENT_SECRET"],
            "scope": "https://graph.microsoft.com/.default",
            "grant_type": "client_credentials",
        },
        timeout=30,
    )
    if antwoord.status_code != 200:
        raise SystemExit(
            "Inloggen bij Microsoft mislukte (status %s). Controleer de secrets "
            "SP_TENANT_ID, SP_CLIENT_ID en SP_CLIENT_SECRET.\n%s"
            % (antwoord.status_code, antwoord.text[:500])
        )
    return antwoord.json()["access_token"]


def graph_get(url, token, **kwargs):
    return requests.get(url, headers={"Authorization": "Bearer " + token}, timeout=60, **kwargs)


def download_uit_sharepoint():
    """Haalt het Excel-bestand op en levert de inhoud als bytes."""
    token = graph_token()

    site_url = "https://graph.microsoft.com/v1.0/sites/%s:%s" % (HOSTNAME, SITE_PATH)
    site = graph_get(site_url, token)
    if site.status_code != 200:
        raise SystemExit(
            "De SharePoint-site %s%s is niet gevonden of de app mag er niet bij "
            "(status %s).\n%s" % (HOSTNAME, SITE_PATH, site.status_code, site.text[:500])
        )
    site_id = site.json()["id"]

    pad = urllib.parse.quote(FILE_PATH)
    item_url = "https://graph.microsoft.com/v1.0/sites/%s/drive/root:/%s" % (site_id, pad)
    item = graph_get(item_url, token)

    if item.status_code == 404:
        # Bestand hernoemd of verplaatst: zoek het alsnog op naam.
        naam = os.path.basename(FILE_PATH)
        zoek = graph_get(
            "https://graph.microsoft.com/v1.0/sites/%s/drive/root/search(q='%s')"
            % (site_id, urllib.parse.quote(os.path.splitext(naam)[0])),
            token,
        )
        treffers = zoek.json().get("value", []) if zoek.status_code == 200 else []
        treffers = [t for t in treffers if t.get("name", "").lower().endswith(".xlsx")]
        if not treffers:
            raise SystemExit(
                "Het bestand '%s' staat niet (meer) op die plek in SharePoint en is "
                "ook niet op naam gevonden. Pas SP_FILE_PATH aan in de workflow." % FILE_PATH
            )
        item = zoek
        gevonden = treffers[0]
        log("Let op: bestand niet op het vaste pad gevonden, maar wel via zoeken: %s"
            % gevonden.get("webUrl", gevonden["name"]))
        item_id = gevonden["id"]
    elif item.status_code != 200:
        raise SystemExit(
            "Ophalen van het bestand mislukte (status %s).\n%s" % (item.status_code, item.text[:500])
        )
    else:
        item_id = item.json()["id"]

    inhoud = graph_get(
        "https://graph.microsoft.com/v1.0/sites/%s/drive/items/%s/content" % (site_id, item_id), token
    )
    if inhoud.status_code != 200:
        raise SystemExit("Downloaden mislukte (status %s)." % inhoud.status_code)
    return inhoud.content


def tekst(waarde):
    """Zet een celwaarde om naar nette tekst (13 in plaats van 13.0)."""
    if waarde is None:
        return ""
    if isinstance(waarde, float) and waarde.is_integer():
        return str(int(waarde))
    if isinstance(waarde, datetime):
        return waarde.strftime("%d-%m-%Y")
    return str(waarde).strip()


def land_uit(waarde):
    schoon = tekst(waarde).lower()
    for code, varianten in LANDEN.items():
        if schoon in varianten or schoon.startswith(code.lower()):
            return code
    return None


def lees_werkblad(ws):
    """
    Zoekt in een werkblad de koprij en levert (leveranciers, gemiste_codes).
    De kolomvolgorde is vrij; kolommen worden herkend aan hun kop.
    """
    rijen = [[cel for cel in rij] for rij in ws.iter_rows(values_only=True)]

    kop_index = None
    for i, rij in enumerate(rijen[:15]):
        koppen = [tekst(c).lower() for c in rij]
        heeft_naam = any(re.search(r"groothandel|leverancier|klant", k) for k in koppen)
        heeft_code = any(re.match(r"^la\s*\d+", k) for k in koppen)
        if heeft_naam and heeft_code:
            kop_index = i
            break

    if kop_index is None:
        return [], []

    koppen = [tekst(c) for c in rijen[kop_index]]
    naam_kolom = next(i for i, k in enumerate(koppen)
                      if re.search(r"groothandel|leverancier|klant", k, re.I))
    land_kolom = next((i for i, k in enumerate(koppen) if re.fullmatch(r"land", k.strip(), re.I)), None)

    code_kolom = {}
    for i, kop in enumerate(koppen):
        treffer = re.match(r"^\s*(LA\s*\d+)", kop, re.I)
        if treffer:
            code_kolom[treffer.group(1).replace(" ", "").upper()] = i

    gemist = [code for code in CODES if code not in code_kolom]

    # Geen Land-kolom? Dan bepaalt de naam van het tabblad het land.
    land_van_blad = land_uit(ws.title)

    leveranciers = []
    for rij in rijen[kop_index + 1:]:
        naam = tekst(rij[naam_kolom]) if naam_kolom < len(rij) else ""
        if not naam:
            continue
        land = None
        if land_kolom is not None and land_kolom < len(rij):
            land = land_uit(rij[land_kolom])
        land = land or land_van_blad
        producten = []
        for code in CODES:
            i = code_kolom.get(code)
            producten.append(tekst(rij[i]) if i is not None and i < len(rij) else "")
        leveranciers.append({"land": land, "name": naam, "products": producten})

    return leveranciers, gemist


def lees_bestand(inhoud):
    wb = load_workbook(io.BytesIO(inhoud), data_only=True, read_only=True)
    alles = []
    gemist = set()
    gelezen_bladen = []

    for ws in wb.worksheets:
        leveranciers, blad_gemist = lees_werkblad(ws)
        if leveranciers:
            gelezen_bladen.append("%s (%d)" % (ws.title, len(leveranciers)))
            alles.extend(leveranciers)
            gemist.update(blad_gemist)

    if not alles:
        raise SystemExit(
            "In het Excel-bestand is geen bruikbare tabel gevonden. Verwacht wordt een "
            "koprij met een kolom 'Groothandel' (of 'Leverancier') en kolommen die "
            "beginnen met een LA-code, bijvoorbeeld 'LA152'."
        )

    log("Gelezen tabbladen: " + ", ".join(gelezen_bladen))
    if gemist:
        log("Let op: geen kolom gevonden voor " + ", ".join(sorted(gemist))
            + " — die blijven op de site leeg.")

    zonder_land = [s["name"] for s in alles if not s["land"]]
    if zonder_land:
        raise SystemExit(
            "Van %d leverancier(s) is het land onbekend, bijvoorbeeld '%s'. Zet in het "
            "Excel-bestand een kolom 'Land' met NL of BE, of geef de tabbladen de naam "
            "'Nederland' en 'Belgie'." % (len(zonder_land), zonder_land[0])
        )

    per_land = {"NL": [], "BE": []}
    for s in alles:
        per_land[s["land"]].append({"name": s["name"], "products": s["products"]})
    return per_land


def aantal_nummers(per_land):
    return sum(1 for lijst in per_land.values() for s in lijst
               for p in s["products"] if p.strip())


def main():
    p = argparse.ArgumentParser(description="SharePoint-lijst omzetten naar leveranciers-data.json")
    p.add_argument("--bestand", help="Lokaal .xlsx-bestand gebruiken in plaats van SharePoint")
    p.add_argument("--uit", default="leveranciers-data.json", help="Doelbestand")
    p.add_argument("--droog", action="store_true", help="Alleen tonen, niets wegschrijven")
    p.add_argument("--minimaal", type=int, default=20,
                   help="Minimaal aantal leveranciers; daaronder stopt het script")
    p.add_argument("--forceer", action="store_true",
                   help="Ook wegschrijven als er veel nummers zouden verdwijnen")
    args = p.parse_args()

    if args.bestand:
        inhoud = open(args.bestand, "rb").read()
        bron = os.path.basename(args.bestand)
    else:
        inhoud = download_uit_sharepoint()
        bron = "%s (SharePoint, site %s)" % (os.path.basename(FILE_PATH), SITE_PATH.strip("/"))

    per_land = lees_bestand(inhoud)
    totaal = len(per_land["NL"]) + len(per_land["BE"])
    nieuw_gevuld = aantal_nummers(per_land)
    log("Gevonden: %d leveranciers (NL %d, BE %d) met %d ingevulde artikelnummers."
        % (totaal, len(per_land["NL"]), len(per_land["BE"]), nieuw_gevuld))

    if totaal < args.minimaal:
        raise SystemExit(
            "Slechts %d leveranciers gevonden (verwacht minstens %d). Het bestand is "
            "waarschijnlijk anders ingedeeld; er wordt niets bijgewerkt."
            % (totaal, args.minimaal)
        )

    # Veiligheidsrem: nooit stilzwijgend een groot deel van de nummers wissen.
    if os.path.exists(args.uit) and not args.forceer:
        oud = json.load(open(args.uit, encoding="utf-8")).get("leveranciers", {})
        oud_gevuld = aantal_nummers({"NL": oud.get("NL", []), "BE": oud.get("BE", [])})
        if oud_gevuld and nieuw_gevuld < oud_gevuld * 0.7:
            raise SystemExit(
                "Er zouden %d van de %d artikelnummers verdwijnen. Dat lijkt op een "
                "wijziging in de indeling van het Excel-bestand, dus er wordt niets "
                "bijgewerkt. Klopt het toch? Draai dan met --forceer."
                % (oud_gevuld - nieuw_gevuld, oud_gevuld)
            )

    data = {
        "bijgewerkt": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
        "bron": bron,
        "codes": CODES,
        "leveranciers": per_land,
    }
    tekst_uit = json.dumps(data, ensure_ascii=False, indent=2, sort_keys=False) + "\n"

    if args.droog:
        log(tekst_uit[:1500])
        log("(--droog: er is niets weggeschreven)")
        return

    open(args.uit, "w", encoding="utf-8").write(tekst_uit)
    log("Weggeschreven naar %s" % args.uit)


if __name__ == "__main__":
    sys.exit(main())
